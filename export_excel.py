"""
export_excel.py — Excel export engine for Multi-SKU AFC Baseline Forecast
Mondelez Vietnam

Sheets produced:
    Overview          — 1 row per SKU: model, params, MAPE, score, FC year, YoY%
    Forecast_Detail   — rows = SKU, cols = months (value Mil VND + YoY% row)
    [SKU_name]        — 1 sheet per SKU: Annual / Quarterly / Monthly growth table
"""

import math
import io
from datetime import datetime
from collections import defaultdict

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Colour palette (matches app theme) ───────────────────────
C_HEADER_BG  = "1E2130"
C_HEADER_FG  = "4A9EFF"
C_HIST_BG    = "1a2640"
C_FC_BG      = "2a1f10"
C_BEST_BG    = "0d2318"
C_BEST_FG    = "10B981"
C_WARN_FG    = "F59400"
C_NEG_FG     = "EF4444"
C_POS_FG     = "10B981"
C_NEU_FG     = "94A3B8"
C_WHITE      = "E2E8F0"
C_SUBTEXT    = "94A3B8"
C_DIVIDER    = "2D3250"
C_SECTION_BG = "131722"

MODEL_LABELS = {
    "trend_seasonal": "Trend + Seasonal",
    "seasonal":       "Seasonal",
    "trend":          "Trend",
    "constant":       "Constant",
}


# ══════════════════════════════════════════════════════════════
# STYLE HELPERS
# ══════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color: str = C_WHITE, bold: bool = False,
          size: int = 10, mono: bool = False) -> Font:
    name = "Courier New" if mono else "Calibri"
    return Font(name=name, color=hex_color, bold=bold, size=size)

def _align(h: str = "center", v: str = "center",
           wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_bottom(color: str = C_DIVIDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def _set(ws, row, col, value, fg=C_WHITE, bg=None, bold=False,
         h="center", size=10, mono=False, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(fg, bold, size, mono)
    cell.alignment = _align(h)
    if bg:
        cell.fill = _fill(bg)
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════════════
# GROWTH CALCULATORS
# ══════════════════════════════════════════════════════════════

def _annual_sums(values, dates):
    d = defaultdict(float)
    for dt, v in zip(dates, values):
        yr = dt.year if isinstance(dt, datetime) else pd.Timestamp(dt).year
        d[yr] += v
    return dict(sorted(d.items()))


def _quarterly_sums(values, dates):
    d = defaultdict(float)
    for dt, v in zip(dates, values):
        ts = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
        q  = (ts.month - 1) // 3 + 1
        d[(ts.year, q)] += v
    return dict(sorted(d.items()))


def _pct(new, old):
    if old and old != 0 and new is not None:
        return new / old - 1
    return None


def _fmt_pct(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return v   # keep as float; apply number format in Excel


def _growth_rows(all_v, all_d, hist_end_date):
    """
    Build Annual / Quarterly / Monthly growth rows.
    Returns list of dicts: {section, label, value, yoy, qoq, mom, is_forecast}
    """
    rows = []
    hist_end = hist_end_date if isinstance(hist_end_date, datetime) \
               else pd.Timestamp(hist_end_date).to_pydatetime()

    # ── Annual ───────────────────────────────────────────────
    ann = _annual_sums(all_v, all_d)
    ann_yrs = sorted(ann)
    for i, yr in enumerate(ann_yrs):
        prev_yr = ann.get(yr - 1)
        rows.append({
            "section":     "Annual",
            "label":       str(yr),
            "value":       ann[yr],
            "yoy":         _pct(ann[yr], prev_yr),
            "qoq":         None,
            "mom":         None,
            "is_forecast": yr > hist_end.year,
        })

    # ── Quarterly ─────────────────────────────────────────────
    qtr = _quarterly_sums(all_v, all_d)
    qtr_keys = sorted(qtr)
    for i, (yr, q) in enumerate(qtr_keys):
        prev_q = qtr.get((yr, q - 1) if q > 1 else (yr - 1, 4))
        prev_yr_q = qtr.get((yr - 1, q))
        last_hist_q = ((hist_end.year, (hist_end.month - 1) // 3 + 1))
        rows.append({
            "section":     "Quarterly",
            "label":       f"Q{q} {yr}",
            "value":       qtr[(yr, q)],
            "yoy":         _pct(qtr[(yr, q)], prev_yr_q),
            "qoq":         _pct(qtr[(yr, q)], prev_q),
            "mom":         None,
            "is_forecast": (yr, q) > last_hist_q,
        })

    # ── Monthly ───────────────────────────────────────────────
    mon_series = sorted(zip(all_d, all_v), key=lambda x: x[0])
    for i, (dt, v) in enumerate(mon_series):
        ts  = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
        prev_m  = mon_series[i - 1][1] if i > 0 else None
        prev_yr_v = next(
            (val for d2, val in mon_series
             if (d2.year if isinstance(d2, datetime) else pd.Timestamp(d2).year) == ts.year - 1
             and (d2.month if isinstance(d2, datetime) else pd.Timestamp(d2).month) == ts.month),
            None
        )
        rows.append({
            "section":     "Monthly",
            "label":       ts.strftime("%b %Y"),
            "value":       v,
            "yoy":         _pct(v, prev_yr_v),
            "qoq":         None,
            "mom":         _pct(v, prev_m) if i > 0 else None,
            "is_forecast": ts > hist_end,
        })

    return rows


# ══════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════

def _write_overview(wb, sku_results, fc_start_yr):
    ws = wb.create_sheet("Overview", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    _set(ws, 1, 1, "AFC BASELINE FORECAST — OVERVIEW",
         fg=C_HEADER_FG, bg=C_SECTION_BG, bold=True, size=12, mono=True, h="left")
    ws.row_dimensions[1].height = 28

    # Sub-header
    ws.merge_cells("A2:L2")
    _set(ws, 2, 1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Mondelez Vietnam / AFC",
         fg=C_SUBTEXT, bg=C_SECTION_BG, size=9, h="left")

    # Column headers
    headers = [
        "PPG", "Channel", "Best Model",
        "α", "β", "γ",
        "MAPE", "Score",
        f"FC {fc_start_yr} (Mil VND)", f"YoY {fc_start_yr}%",
        "YoY Anchor%", "Status",
    ]
    for ci, h in enumerate(headers, 1):
        _set(ws, 4, ci, h, fg=C_HEADER_FG, bg=C_HEADER_BG, bold=True, size=9, mono=True)
    ws.row_dimensions[4].height = 20

    # Data rows
    for ri, sr in enumerate(sku_results, 5):
        best = sr["best"]
        sb   = best.get("score_breakdown", {})
        fc_yoy = sb.get("yoy_forecast", {}).get(fc_start_yr)
        anchor = sb.get("yoy_anchor", {}).get(fc_start_yr)
        fc_yr_vol = sr.get("fc_year_vol", {}).get(fc_start_yr)

        # Determine status
        if fc_yoy is not None and anchor is not None:
            diff = abs(fc_yoy/100 - anchor/100)
            status = "✅ On track" if diff < 0.05 else ("⚠️ Review" if diff < 0.15 else "❌ Off anchor")
        else:
            status = "—"

        bg = C_SECTION_BG if ri % 2 == 0 else C_HEADER_BG
        row_data = [
            sr["ppg"], sr["channel"],
            MODEL_LABELS.get(best["model"], best["model"]),
            best["alpha"], best["beta"], best["gamma"],
            best["mape"], round(best["score"], 4),
            fc_yr_vol, fc_yoy,
            anchor, status,
        ]
        for ci, v in enumerate(row_data, 1):
            cell = _set(ws, ri, ci, v, bg=bg, h="center")
            # Format specific columns
            if ci == 7:   # MAPE
                cell.number_format = "0.0%"
            elif ci in (9,):   # FC volume
                cell.number_format = '#,##0.00'
            elif ci in (10, 11):  # YoY%
                if v is not None:
                    try:
                        cell.value = float(v) / 100
                        cell.number_format = '+0.0%;-0.0%;0.0%'
                        fg = C_POS_FG if float(v) > 0 else (C_NEG_FG if float(v) < 0 else C_NEU_FG)
                        cell.font = _font(fg, bold=True, size=10)
                    except (TypeError, ValueError):
                        pass

        ws.row_dimensions[ri].height = 18

    # Column widths
    widths = [28, 10, 18, 7, 7, 7, 8, 9, 18, 12, 13, 12]
    for ci, w in enumerate(widths, 1):
        _col_width(ws, ci, w)

    ws.freeze_panes = "A5"


def _write_forecast_detail(wb, sku_results, all_months):
    ws = wb.create_sheet("Forecast_Detail")
    ws.sheet_view.showGridLines = False

    # Title
    n_cols = 4 + len(all_months)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    _set(ws, 1, 1, "FORECAST DETAIL — ALL SKUs (Mil VND)",
         fg=C_HEADER_FG, bg=C_SECTION_BG, bold=True, size=12, mono=True, h="left")
    ws.row_dimensions[1].height = 28

    # Column headers: SKU info + months
    hdr_row = 3
    fixed_hdrs = ["PPG", "Channel", "Best Model", "Row Type"]
    for ci, h in enumerate(fixed_hdrs, 1):
        _set(ws, hdr_row, ci, h, fg=C_HEADER_FG, bg=C_HEADER_BG, bold=True, size=9, mono=True)
    for ci, dt in enumerate(all_months, 5):
        label = dt.strftime("%b %Y") if isinstance(dt, datetime) else pd.Timestamp(dt).strftime("%b %Y")
        _set(ws, hdr_row, ci, label, fg=C_HEADER_FG, bg=C_HEADER_BG, bold=True, size=8, mono=True)

    ws.row_dimensions[hdr_row].height = 20

    # Mark forecast boundary
    # Find first forecast month col
    def _is_fc_month(dt, hist_end):
        ts = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
        he = hist_end if isinstance(hist_end, datetime) else pd.Timestamp(hist_end).to_pydatetime()
        return ts > he

    ri = hdr_row + 1
    for sr in sku_results:
        best       = sr["best"]
        hist_v     = sr["history"]
        hist_d     = sr["hist_dates"]
        fc_v       = best["forecast"]
        fc_d       = sr["fc_dates"]
        hist_end   = hist_d[-1]

        # Build month → value map
        all_v_map = {}
        for dt, v in zip(hist_d, hist_v):
            k = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
            all_v_map[k.replace(day=1)] = ("hist", v)
        for dt, v in zip(fc_d, fc_v):
            k = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
            all_v_map[k.replace(day=1)] = ("fc", v)

        # Build YoY map
        yoy_map = {}
        for dt, (typ, v) in all_v_map.items():
            prev_yr_dt = dt.replace(year=dt.year - 1)
            if prev_yr_dt in all_v_map:
                prev_v = all_v_map[prev_yr_dt][1]
                yoy_map[dt] = v / prev_v - 1 if prev_v else None

        model_lbl = MODEL_LABELS.get(best["model"], best["model"])
        bg_even   = C_SECTION_BG

        # Row 1: values (Mil VND)
        _set(ws, ri, 1, sr["ppg"],     bg=bg_even, h="left")
        _set(ws, ri, 2, sr["channel"], bg=bg_even)
        _set(ws, ri, 3, model_lbl,     bg=bg_even)
        _set(ws, ri, 4, "Value (Mil VND)", fg=C_SUBTEXT, bg=bg_even, size=9)
        for ci, dt in enumerate(all_months, 5):
            k  = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
            k  = k.replace(day=1)
            entry = all_v_map.get(k)
            if entry:
                typ, v = entry
                bg = C_FC_BG if typ == "fc" else C_HIST_BG
                cell = _set(ws, ri, ci, round(v / 1e6, 4) if v else 0,
                            bg=bg, num_fmt='#,##0.0000')
                cell.font = _font(C_WHITE if typ == "hist" else C_WARN_FG, size=9)
        ws.row_dimensions[ri].height = 16
        ri += 1

        # Row 2: YoY%
        _set(ws, ri, 1, "", bg=bg_even)
        _set(ws, ri, 2, "", bg=bg_even)
        _set(ws, ri, 3, "", bg=bg_even)
        _set(ws, ri, 4, "YoY%", fg=C_SUBTEXT, bg=bg_even, size=9)
        for ci, dt in enumerate(all_months, 5):
            k = dt if isinstance(dt, datetime) else pd.Timestamp(dt).to_pydatetime()
            k = k.replace(day=1)
            v = yoy_map.get(k)
            if v is not None:
                typ = all_v_map.get(k, ("hist",))[0]
                bg  = C_FC_BG if typ == "fc" else C_HIST_BG
                fg  = C_POS_FG if v > 0.005 else (C_NEG_FG if v < -0.005 else C_NEU_FG)
                cell = _set(ws, ri, ci, v, bg=bg, num_fmt='+0.0%;-0.0%;0.0%')
                cell.font = _font(fg, bold=True, size=9)
        ws.row_dimensions[ri].height = 16
        ri += 1

        # Blank separator row
        for ci in range(1, 5 + len(all_months)):
            ws.cell(row=ri, column=ci).fill = _fill(C_SECTION_BG)
        ws.row_dimensions[ri].height = 4
        ri += 1

    # Column widths
    _col_width(ws, 1, 28)
    _col_width(ws, 2, 10)
    _col_width(ws, 3, 18)
    _col_width(ws, 4, 16)
    for ci in range(5, 5 + len(all_months)):
        _col_width(ws, ci, 11)

    ws.freeze_panes = "E4"


def _write_sku_growth(wb, sr):
    # Sheet name: max 31 chars, no special chars
    raw_name = f"{sr['ppg']}_{sr['channel']}"
    sheet_name = raw_name.replace("*", "x").replace("/", "-").replace(":", "")[:31]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    best     = sr["best"]
    hist_v   = sr["history"]
    hist_d   = sr["hist_dates"]
    fc_v     = best["forecast"]
    fc_d     = sr["fc_dates"]
    hist_end = hist_d[-1]

    all_v = hist_v + list(fc_v)
    all_d = list(hist_d) + list(fc_d)

    # Title
    ws.merge_cells("A1:H1")
    _set(ws, 1, 1, f"{sr['ppg']} | {sr['channel']} — Growth Analysis",
         fg=C_HEADER_FG, bg=C_SECTION_BG, bold=True, size=11, mono=True, h="left")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    model_lbl = MODEL_LABELS.get(best["model"], best["model"])
    _set(ws, 2, 1,
         f"Model: {model_lbl}  |  α={best['alpha']:.2f}  β={best['beta']:.2f}  "
         f"γ={best['gamma']:.2f}  |  MAPE={best['mape']*100:.1f}%  |  Score={best['score']:.4f}",
         fg=C_SUBTEXT, bg=C_SECTION_BG, size=9, h="left")
    ws.row_dimensions[2].height = 16

    # Column headers
    cols = ["Section", "Period", "Value (Mil VND)", "YoY%", "QoQ%", "MoM%", "Type", "Note"]
    for ci, h in enumerate(cols, 1):
        _set(ws, 4, ci, h, fg=C_HEADER_FG, bg=C_HEADER_BG, bold=True, size=9, mono=True)
    ws.row_dimensions[4].height = 20

    # Growth rows
    growth_rows = _growth_rows(all_v, all_d, hist_end)

    prev_section = None
    ri = 5
    for gr in growth_rows:
        section     = gr["section"]
        is_forecast = gr["is_forecast"]
        bg = C_FC_BG if is_forecast else C_HIST_BG

        # Section separator
        if section != prev_section:
            ws.merge_cells(f"A{ri}:H{ri}")
            _set(ws, ri, 1, f"── {section.upper()} ──",
                 fg=C_HEADER_FG, bg=C_SECTION_BG, bold=True, size=9, mono=True, h="left")
            ws.row_dimensions[ri].height = 16
            ri += 1
            prev_section = section

        # Data row
        row_vals = [
            section,
            gr["label"],
            round(gr["value"] / 1e6, 4) if gr["value"] else 0,
            gr["yoy"],
            gr["qoq"],
            gr["mom"],
            "Forecast" if is_forecast else "History",
            "",
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = _set(ws, ri, ci, v, bg=bg, h="center" if ci != 2 else "left")
            cell.font = _font(
                C_WARN_FG if is_forecast else C_WHITE,
                size=9
            )
            # Number formats
            if ci == 3:  # value
                cell.number_format = '#,##0.0000'
            elif ci in (4, 5, 6):  # pct
                if v is not None:
                    cell.value       = float(v)
                    cell.number_format = '+0.0%;-0.0%;0.0%'
                    fg = C_POS_FG if float(v) > 0.005 else \
                         (C_NEG_FG if float(v) < -0.005 else C_NEU_FG)
                    cell.font = _font(fg, bold=True, size=9)
        ws.row_dimensions[ri].height = 16
        ri += 1

    # Widths
    widths = [12, 14, 18, 12, 12, 12, 10, 14]
    for ci, w in enumerate(widths, 1):
        _col_width(ws, ci, w)

    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════
# MAIN EXPORT FUNCTION
# ══════════════════════════════════════════════════════════════

def export_multi_sku(sku_results: list, hist_end_date) -> bytes:
    """
    Build full Excel workbook for multi-SKU results.

    sku_results: list of {
        ppg, channel, name,
        history, hist_dates, fc_dates,
        best          : result dict from select_best
        fc_year_vol   : {year: total_volume}
    }

    Returns bytes (Excel file content).
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required. Run: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # Determine forecast start year
    all_fc_dates = []
    for sr in sku_results:
        all_fc_dates.extend(sr["fc_dates"])
    fc_start_yr = min(
        d.year if isinstance(d, datetime) else pd.Timestamp(d).year
        for d in all_fc_dates
    ) if all_fc_dates else datetime.now().year + 1

    # Build full month list (union of all hist + fc months)
    all_month_set = set()
    for sr in sku_results:
        for d in list(sr["hist_dates"]) + list(sr["fc_dates"]):
            k = d if isinstance(d, datetime) else pd.Timestamp(d).to_pydatetime()
            all_month_set.add(k.replace(day=1))
    all_months = sorted(all_month_set)

    # ── Sheet 1: Overview ─────────────────────────────────────
    _write_overview(wb, sku_results, fc_start_yr)

    # ── Sheet 2: Forecast Detail ──────────────────────────────
    _write_forecast_detail(wb, sku_results, all_months)

    # ── Sheet per SKU ─────────────────────────────────────────
    for sr in sku_results:
        _write_sku_growth(wb, sr)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
