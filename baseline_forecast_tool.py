"""
baseline_forecast_tool.py — Forecast Engine
Mondelez Vietnam / AFC Baseline Forecast App

Exports:
    read_single_sku_xlsm(path)       → skus, file_params
    read_multi_sku_excel(path)        → skus, file_params
    read_csv(path)                    → skus, file_params
    estimate_grid_size(params, step)  → dict
    grid_search(history, params, h_steps, period, step, verbose) → list[result]
    select_best(results, history, top_n) → list[result]  sorted by business score
    export_to_excel(path, all_sku_results)
    _seasonality_strength(history, period)
    _yoy_growth(values, period)
    _qoq_growth(values)
    _mom_growth(values)
    _var(seq)
"""

import math
import warnings
from collections import defaultdict
from datetime import datetime
from itertools import product

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════
# INTERNAL MATH HELPERS
# ══════════════════════════════════════════════════════════════

def _var(seq):
    """Variance of a list, ignoring None / nan."""
    vals = [v for v in seq if v is not None and not (isinstance(v, float) and math.isnan(v))]
    if len(vals) < 2:
        return 0.0
    m = sum(vals) / len(vals)
    return sum((x - m) ** 2 for x in vals) / len(vals)


def _seasonality_strength(history, period: int = 12) -> float:
    """
    Ratio-to-moving-average seasonality strength (0–1).
    Higher = stronger seasonal pattern.
    """
    arr = np.array(history, float)
    n = len(arr)
    if n < period * 2:
        return 0.0
    half = period // 2
    # centred MA
    trend = np.full(n, np.nan)
    for i in range(half, n - half):
        trend[i] = np.mean(arr[i - half: i + half + 1])
    # seasonal ratios per position
    by_pos = defaultdict(list)
    for i in range(n):
        if not np.isnan(trend[i]) and trend[i] > 0:
            by_pos[i % period].append(arr[i] / trend[i])
    if not by_pos:
        return 0.0
    s_factors = np.array([np.mean(v) for v in by_pos.values()])
    # strength = std of seasonal factors / (std + residual proxy)
    std_s = float(np.std(s_factors))
    detrended = arr[~np.isnan(trend)] / trend[~np.isnan(trend)]
    std_total = float(np.std(detrended)) if len(detrended) > 1 else 1.0
    if std_total == 0:
        return 0.0
    return min(1.0, std_s / std_total)


def _yoy_growth(values, period: int = 12):
    """List of YoY growth rates (values[i] / values[i-period] - 1)."""
    out = []
    for i in range(period, len(values)):
        prev = values[i - period]
        if prev and prev != 0:
            out.append(values[i] / prev - 1)
    return out


def _qoq_growth(values):
    """Approximate QoQ growth using 3-month sums."""
    if len(values) < 6:
        return []
    qtrs = [sum(values[i: i + 3]) for i in range(0, len(values) - 2, 3)]
    out = []
    for i in range(1, len(qtrs)):
        if qtrs[i - 1] and qtrs[i - 1] != 0:
            out.append(qtrs[i] / qtrs[i - 1] - 1)
    return out


def _mom_growth(values):
    """MoM growth rates."""
    out = []
    for i in range(1, len(values)):
        if values[i - 1] and values[i - 1] != 0:
            out.append(values[i] / values[i - 1] - 1)
    return out


# ══════════════════════════════════════════════════════════════
# HOLT-WINTERS MODELS
# ══════════════════════════════════════════════════════════════

def _hw_trend_seasonal(history, alpha, beta, gamma, h, period=12):
    """
    Multiplicative Holt-Winters (Trend + Seasonal).
    Returns forecast list and fitted (expost) list.
    """
    n = len(history)
    if n < period * 2:
        return [history[-1]] * h, []

    # Initialise
    level = np.mean(history[:period])
    trend = (np.mean(history[period: period * 2]) - np.mean(history[:period])) / period
    seasonal = [history[i] / level for i in range(period)]

    fitted = []
    for t in range(n):
        s_idx = t % period
        prev_level = level
        prev_trend = trend
        obs = history[t]
        # update
        level = alpha * (obs / seasonal[s_idx]) + (1 - alpha) * (prev_level + prev_trend)
        trend = beta * (level - prev_level) + (1 - beta) * prev_trend
        seasonal[s_idx] = gamma * (obs / level) + (1 - gamma) * seasonal[s_idx]
        fitted.append((prev_level + prev_trend) * seasonal[s_idx])

    forecast = []
    for i in range(1, h + 1):
        s_idx = (n + i - 1) % period
        forecast.append((level + trend * i) * seasonal[s_idx])

    return forecast, fitted


def _hw_seasonal(history, alpha, gamma, h, period=12):
    """
    Seasonal-only Holt-Winters (no trend component).
    Returns forecast list and fitted list.
    """
    n = len(history)
    if n < period * 2:
        return [history[-1]] * h, []

    level = np.mean(history[:period])
    seasonal = [history[i] / level for i in range(period)]

    fitted = []
    for t in range(n):
        s_idx = t % period
        prev_level = level
        obs = history[t]
        level = alpha * (obs / seasonal[s_idx]) + (1 - alpha) * prev_level
        seasonal[s_idx] = gamma * (obs / level) + (1 - gamma) * seasonal[s_idx]
        fitted.append(prev_level * seasonal[s_idx])

    forecast = []
    for i in range(1, h + 1):
        s_idx = (n + i - 1) % period
        forecast.append(level * seasonal[s_idx])

    return forecast, fitted


def _hw_trend(history, alpha, beta, h):
    """
    Holt's double exponential (Trend only, no seasonal).
    Returns forecast list and fitted list.
    """
    n = len(history)
    level = history[0]
    trend = history[1] - history[0] if n > 1 else 0.0

    fitted = []
    for t in range(n):
        obs = history[t]
        prev_level = level
        prev_trend = trend
        level = alpha * obs + (1 - alpha) * (prev_level + prev_trend)
        trend = beta * (level - prev_level) + (1 - beta) * prev_trend
        fitted.append(prev_level + prev_trend)

    forecast = [level + trend * i for i in range(1, h + 1)]
    return forecast, fitted


def _hw_constant(history, alpha, h):
    """
    Simple exponential smoothing (constant / flat).
    Returns forecast list and fitted list.
    """
    level = history[0]
    fitted = []
    for obs in history:
        fitted.append(level)
        level = alpha * obs + (1 - alpha) * level

    forecast = [level] * h
    return forecast, fitted


# ══════════════════════════════════════════════════════════════
# METRICS
# ══════════════════════════════════════════════════════════════

def _mape(actual, predicted):
    """MAPE — skips zeros in actual."""
    errors = []
    for a, p in zip(actual, predicted):
        if a and a != 0:
            errors.append(abs(a - p) / abs(a))
    return float(np.mean(errors)) if errors else float("nan")


def _annual_vol(values: list, dates: list) -> dict:
    """Sum monthly values by calendar year → {year: total_vol}."""
    d = defaultdict(float)
    for dt, v in zip(dates, values):
        if isinstance(dt, datetime):
            d[dt.year] += v
        else:
            try:
                d[pd.Timestamp(dt).year] += v
            except Exception:
                pass
    return dict(sorted(d.items()))


def _quarterly_vol(values: list, dates: list) -> dict:
    """Sum monthly values by (year, quarter) → {(year,q): total_vol}."""
    d = defaultdict(float)
    for dt, v in zip(dates, values):
        if isinstance(dt, datetime):
            q = (dt.month - 1) // 3 + 1
            d[(dt.year, q)] += v
        else:
            try:
                ts = pd.Timestamp(dt)
                q  = (ts.month - 1) // 3 + 1
                d[(ts.year, q)] += v
            except Exception:
                pass
    return dict(sorted(d.items()))


def _yoy_series(ann: dict) -> dict:
    """YoY growth dict: {year: growth_rate} — needs ≥2 years."""
    yrs = sorted(ann)
    out = {}
    for i in range(1, len(yrs)):
        y, yp = yrs[i], yrs[i-1]
        if ann[yp] and ann[yp] != 0:
            out[y] = ann[y] / ann[yp] - 1
    return out


def _extrapolate_anchor(yoy_hist: dict, full_ann: dict = None) -> dict:
    """
    Extrapolate expected YoY anchor for each forecast year.

    Uses only FULL calendar years (12 months of data) to avoid partial-year
    distortion (e.g. Jan-Apr only skewing the calculation).

    Logic:
      - Filter to full years only (proxy: vol >= 0.65 x mean annual vol)
      - Fit linear regression on YoY series of full years
      - Special rules:
          Last 2 full-year YoY both negative -> anchor = 0.0 (expect flat recovery)
          Last 2 full-year YoY both positive -> continue slope
          Mixed                              -> slope dampened x0.7
      - Cap: anchors clipped to [-0.40, +0.40]

    Returns dict {forecast_year: anchor_yoy_rate}
    """
    if not yoy_hist:
        return {}

    # Exclude partial years using annual volume as proxy
    if full_ann and len(full_ann) >= 3:
        mean_vol  = np.mean(list(full_ann.values()))
        full_yrs  = {y for y, v in full_ann.items() if v >= 0.65 * mean_vol}
        yoy_clean = {y: v for y, v in yoy_hist.items()
                     if y in full_yrs and (y - 1) in full_yrs}
    else:
        yoy_clean = dict(yoy_hist)

    if not yoy_clean:
        yoy_clean = dict(yoy_hist)

    yrs     = sorted(yoy_clean)
    vals    = [yoy_clean[y] for y in yrs]
    last_yr = yrs[-1]

    if len(vals) >= 2:
        x     = np.arange(len(vals), dtype=float)
        slope = float(np.polyfit(x, vals, 1)[0])
    else:
        slope = 0.0

    last2 = vals[-2:] if len(vals) >= 2 else vals

    anchors = {}
    for offset in range(1, 3):
        fc_yr = last_yr + offset
        if all(v < 0 for v in last2):
            a = 0.0 if offset == 1 else max(0.0, vals[-1] * 0.3)
        elif all(v > 0 for v in last2):
            a = vals[-1] + slope * offset
        else:
            a = vals[-1] + slope * offset * 0.7
        anchors[fc_yr] = float(np.clip(a, -0.40, 0.40))

    return anchors


def _quarterly_cross_year_penalty(
    full_hist_v: list, full_hist_d: list,
    forecast: list,    fc_dates: list,
    target_years: list,
) -> float:
    """
    For each quarter position (Q1/Q2/Q3/Q4), compare YoY growth across years:

        Q1: hist_yr1 → hist_yr2 → hist_yr3 → fc_yr1 → fc_yr2
        Q2: ...
        ...

    Compute the linear trend of each Q's YoY series from history,
    then penalise if forecast deviates from that trend.

    Returns mean absolute deviation (lower = more consistent quarterly trend).
    """
    # Build quarterly volumes for full series
    all_v = list(full_hist_v) + list(forecast)
    all_d = list(full_hist_d) + list(fc_dates)
    qvol  = _quarterly_vol(all_v, all_d)

    # For each quarter position, collect (year, volume) series
    by_q = defaultdict(dict)   # {q: {year: vol}}
    for (yr, q), v in qvol.items():
        by_q[q][yr] = v

    penalties = []
    for q in range(1, 5):
        if q not in by_q:
            continue
        q_series = by_q[q]
        yrs_sorted = sorted(q_series)
        if len(yrs_sorted) < 3:
            continue

        # Compute YoY for this quarter across years
        q_yoy = {}
        for i in range(1, len(yrs_sorted)):
            y, yp = yrs_sorted[i], yrs_sorted[i-1]
            if q_series[yp] and q_series[yp] > 0:
                q_yoy[y] = q_series[y] / q_series[yp] - 1

        if len(q_yoy) < 2:
            continue

        q_yoy_yrs  = sorted(q_yoy)
        q_yoy_vals = [q_yoy[y] for y in q_yoy_yrs]

        # Fit trend on history portion only
        hist_q_yoy = {y: v for y, v in q_yoy.items()
                      if y not in target_years}
        if len(hist_q_yoy) < 2:
            continue

        hx = np.arange(len(hist_q_yoy), dtype=float)
        hy = list(hist_q_yoy.values())
        slope = float(np.polyfit(hx, hy, 1)[0]) if len(hx) >= 2 else 0.0
        last_hist_val = hy[-1]

        # Penalty for each forecast year's quarter
        for offset, yr in enumerate(target_years, 1):
            if yr in q_yoy:
                expected = last_hist_val + slope * offset
                expected = float(np.clip(expected, -0.30, 0.30))
                penalties.append(abs(q_yoy[yr] - expected))

    return float(np.mean(penalties)) if penalties else 0.0


def _business_score(
    history:          list,
    hist_dates:       list,
    forecast:         list,
    fc_dates:         list,
    yoy_anchor_extra: list = None,   # [(date, vol)] Jan-Apr of base year for full Y calc
    period:           int  = 12,
) -> dict:
    """
    Mondelez AFC business scoring — growth-first.

    Scoring priority:
      1. Annual YoY direction vs extrapolated anchor   weight 50
         (does forecast continue the historical growth trend?)
      2. Quarterly cross-year trend consistency        weight 30
         (does Q1/Q2/Q3/Q4 YoY each follow their own trend?)
      3. YoY variance across all years                 weight 10
      4. MAPE ex-post accuracy                         weight  5
      5. MoM variance (smoothness)                     weight  5

    Lower score = better.
    Returns dict with score + full breakdown.
    """
    # ── Prepend extra months (e.g. Jan-Apr 2023) for full Y2023 ──
    extra_v, extra_d = [], []
    if yoy_anchor_extra:
        for d, v in yoy_anchor_extra:
            extra_v.append(float(v))
            extra_d.append(d)

    full_hist_v = extra_v + list(history)
    full_hist_d = extra_d + list(hist_dates)

    all_v = full_hist_v + list(forecast)
    all_d = full_hist_d + list(fc_dates)

    # ── Annual volumes & YoY ─────────────────────────────────
    ann_hist = _annual_vol(full_hist_v, full_hist_d)
    ann_all  = _annual_vol(all_v, all_d)
    yoy_hist = _yoy_series(ann_hist)
    yoy_all  = _yoy_series(ann_all)

    # Forecast years to evaluate: first fc year → 2027 (skip 2028+)
    def _yr(d):
        return d.year if isinstance(d, datetime) else pd.Timestamp(d).year
    fc_start_yr = _yr(min(fc_dates))
    target_years = sorted([y for y in ann_all
                           if fc_start_yr <= y <= fc_start_yr + 1])  # Y+1 and Y+2 only

    # ── 1. YoY anchor penalty ────────────────────────────────
    anchors = _extrapolate_anchor(yoy_hist, full_ann=ann_hist)   # {yr: expected_yoy}
    yoy_anchor_penalty = 0.0
    n_pen = 0
    for yr in target_years:
        if yr in yoy_all and yr in anchors:
            yoy_anchor_penalty += abs(yoy_all[yr] - anchors[yr]) ** 2
            n_pen += 1
    if n_pen:
        yoy_anchor_penalty = math.sqrt(yoy_anchor_penalty / n_pen)   # RMSE

    # ── 2. Quarterly cross-year trend penalty ────────────────
    qoq_trend_pen = _quarterly_cross_year_penalty(
        full_hist_v, full_hist_d, forecast, fc_dates, target_years
    )

    # ── 3. YoY variance (all years incl. forecast) ───────────
    yoy_vals = list(yoy_all.values())
    var_yoy  = float(np.var(yoy_vals)) if len(yoy_vals) >= 2 else 0.0

    # ── 4. MAPE ──────────────────────────────────────────────
    mape_val = _mape(history[-period:], forecast[:period]) \
               if len(forecast) >= period else float("nan")
    mape_val = mape_val if not math.isnan(mape_val) else 1.0

    # ── 5. MoM variance ──────────────────────────────────────
    var_mom = _var(_mom_growth(list(forecast)))

    # ── Composite score ──────────────────────────────────────
    score = (
        50 * yoy_anchor_penalty +
        30 * qoq_trend_pen      +
        10 * var_yoy            +
         5 * mape_val           +
         5 * var_mom
    )

    return {
        "score":              float(score),
        "yoy_anchor":         {y: round(v * 100, 2) for y, v in anchors.items()},
        "yoy_forecast":       {y: round(v * 100, 2) for y, v in yoy_all.items()
                               if y >= min(ann_hist, default=2020)
                               and y <= fc_start_yr + 1},
        "yoy_anchor_penalty": round(yoy_anchor_penalty, 5),
        "qoq_trend_penalty":  round(qoq_trend_pen, 5),
        "var_yoy":            round(var_yoy, 5),
        "mape":               round(mape_val, 5),
        "var_mom":            round(var_mom, 5),
    }


# ══════════════════════════════════════════════════════════════
# GRID SIZE ESTIMATOR
# ══════════════════════════════════════════════════════════════

def estimate_grid_size(params: dict, step: float) -> dict:
    """
    Returns count of grid combinations per model type and total.
    """
    def _n(lo, hi, s):
        return max(1, round((hi - lo) / s) + 1)

    na = _n(params["alpha_min"], params["alpha_max"], step)
    nb = _n(params["beta_min"],  params["beta_max"],  step)
    ng = _n(params["gamma_min"], params["gamma_max"], step)

    ts = na * nb * ng
    s  = na * ng
    t  = na * nb
    c  = na
    return {
        "trend_seasonal": ts,
        "seasonal": s,
        "trend": t,
        "constant": c,
        "total": ts + s + t + c,
    }


# ══════════════════════════════════════════════════════════════
# GRID SEARCH
# ══════════════════════════════════════════════════════════════

def _arange(lo, hi, step):
    vals = []
    v = lo
    while v <= hi + 1e-9:
        vals.append(round(v, 4))
        v += step
    return vals


def grid_search(
    history:            list,
    params:             dict,
    h_steps:            int  = 24,
    period:             int  = 12,
    step:               float = 0.05,
    hist_dates:         list = None,
    yoy_anchor_extra:   list = None,
    verbose:            bool = False,
) -> list:
    """
    Run grid search over α, β, γ for all 4 model types.

    Args:
        history           : monthly volumes (36M, May 2023 → Apr 2026)
        params            : alpha/beta/gamma min/max dict
        h_steps           : forecast horizon (default 24 = May 2026 → Apr 2028)
        period            : seasonality period (12)
        step              : grid step size
        hist_dates        : list of datetime for history months
        yoy_anchor_extra  : [(datetime, vol)] for Jan-Apr 2023 to complete Y2023
        verbose           : print progress

    Returns list of result dicts:
        {
          model, alpha, beta, gamma,
          forecast, expost, mape, score,
          score_breakdown: {yoy_anchor, yoy_anchor_penalty,
                            qoq_consistency, var_yoy, mape, var_mom,
                            yoy_forecast}
        }
    """
    from dateutil.relativedelta import relativedelta as _rd

    hist = [float(v) for v in history]
    n    = len(hist)

    # Build hist_dates if not provided (assume monthly from Jan 2020 — fallback)
    if hist_dates is None:
        hist_dates = [datetime(2023, 5, 1) + _rd(months=i) for i in range(n)]

    # Build fc_dates
    fc_dates = [hist_dates[-1] + _rd(months=i+1) for i in range(h_steps)]

    alphas = _arange(params["alpha_min"], params["alpha_max"], step)
    betas  = _arange(params["beta_min"],  params["beta_max"],  step)
    gammas = _arange(params["gamma_min"], params["gamma_max"], step)

    def _score(fc):
        sd = _business_score(
            hist, hist_dates, fc, fc_dates,
            yoy_anchor_extra=yoy_anchor_extra,
            period=period,
        )
        return sd

    results = []

    # ── Trend + Seasonal ──────────────────────────────────────
    for a, b, g in product(alphas, betas, gammas):
        try:
            fc, ep = _hw_trend_seasonal(hist, a, b, g, h_steps, period)
            sd = _score(fc)
            m  = _mape(hist, ep) if ep else float("nan")
            results.append(dict(model="trend_seasonal", alpha=a, beta=b, gamma=g,
                                forecast=fc, expost=ep, mape=m,
                                score=sd["score"], score_breakdown=sd))
        except Exception:
            pass

    # ── Seasonal ──────────────────────────────────────────────
    for a, g in product(alphas, gammas):
        try:
            fc, ep = _hw_seasonal(hist, a, g, h_steps, period)
            sd = _score(fc)
            m  = _mape(hist, ep) if ep else float("nan")
            results.append(dict(model="seasonal", alpha=a, beta=0.0, gamma=g,
                                forecast=fc, expost=ep, mape=m,
                                score=sd["score"], score_breakdown=sd))
        except Exception:
            pass

    # ── Trend ─────────────────────────────────────────────────
    for a, b in product(alphas, betas):
        try:
            fc, ep = _hw_trend(hist, a, b, h_steps)
            sd = _score(fc)
            m  = _mape(hist, ep) if ep else float("nan")
            results.append(dict(model="trend", alpha=a, beta=b, gamma=0.0,
                                forecast=fc, expost=ep, mape=m,
                                score=sd["score"], score_breakdown=sd))
        except Exception:
            pass

    # ── Constant ──────────────────────────────────────────────
    for a in alphas:
        try:
            fc, ep = _hw_constant(hist, a, h_steps)
            sd = _score(fc)
            m  = _mape(hist, ep) if ep else float("nan")
            results.append(dict(model="constant", alpha=a, beta=0.0, gamma=0.0,
                                forecast=fc, expost=ep, mape=m,
                                score=sd["score"], score_breakdown=sd))
        except Exception:
            pass

    if verbose:
        print(f"[grid_search] {len(results)} combinations evaluated.")
    return results


# ══════════════════════════════════════════════════════════════
# SELECT BEST
# ══════════════════════════════════════════════════════════════

def select_best(results: list, history: list = None, top_n: int = 10) -> list:
    """
    Sort results by composite business score (ascending), return top_n.
    Primary: YoY growth direction vs anchor.
    Secondary: QoQ consistency, YoY variance, MAPE, MoM variance.
    """
    valid = [r for r in results if r.get("score") is not None
             and not math.isnan(float(r["score"])) and not math.isinf(float(r["score"]))]
    valid.sort(key=lambda r: float(r["score"]))
    return valid[:top_n]


# ══════════════════════════════════════════════════════════════
# FILE READERS
# ══════════════════════════════════════════════════════════════

def _parse_date_col(series) -> list:
    """
    Try to parse a date column → list of datetime objects.
    Handles: datetime, 'YYYY-MM', 'YYYY-MM-DD', period strings.
    """
    dates = []
    for val in series:
        if isinstance(val, (pd.Timestamp, datetime)):
            dates.append(pd.Timestamp(val).to_pydatetime().replace(day=1))
        elif isinstance(val, str):
            for fmt in ("%Y-%m", "%Y-%m-%d", "%m/%Y", "%d/%m/%Y"):
                try:
                    dates.append(datetime.strptime(val.strip()[:10], fmt).replace(day=1))
                    break
                except ValueError:
                    pass
            else:
                try:
                    dates.append(pd.to_datetime(val).to_pydatetime().replace(day=1))
                except Exception:
                    dates.append(None)
        else:
            try:
                dates.append(pd.to_datetime(val).to_pydatetime().replace(day=1))
            except Exception:
                dates.append(None)
    return dates


def _default_file_params():
    return {
        "alpha_min": 0.01, "alpha_max": 0.30,
        "beta_min":  0.01, "beta_max":  0.30,
        "gamma_min": 0.01, "gamma_max": 0.99,
    }


def read_csv(path: str):
    """
    Read CSV with columns: date (YYYY-MM), qty [, sku].
    Returns (skus, file_params).
    skus = list of { name, history, dates }
    """
    df = pd.read_csv(path)
    df.columns = [c.lower().strip() for c in df.columns]

    # Normalise column names
    date_col = next((c for c in df.columns if "date" in c or "period" in c or "month" in c), None)
    qty_col  = next((c for c in df.columns if c in ("qty", "quantity", "volume", "sales", "value")), None)
    sku_col  = next((c for c in df.columns if "sku" in c or "material" in c or "product" in c), None)

    if date_col is None or qty_col is None:
        raise ValueError(
            f"CSV must have date and qty columns. Found: {list(df.columns)}"
        )

    df[date_col] = _parse_date_col(df[date_col])
    df.dropna(subset=[date_col, qty_col], inplace=True)
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)

    if sku_col:
        skus_out = []
        for sku_name, grp in df.groupby(sku_col):
            grp = grp.sort_values(date_col)
            skus_out.append({
                "name":    str(sku_name),
                "history": grp[qty_col].tolist(),
                "dates":   grp[date_col].tolist(),
            })
        return skus_out, _default_file_params()
    else:
        df = df.sort_values(date_col)
        return [{
            "name":    "SKU",
            "history": df[qty_col].tolist(),
            "dates":   df[date_col].tolist(),
        }], _default_file_params()


def read_single_sku_xlsm(path: str):
    """
    Read single-SKU xlsm (Baseline_Tool format).
    Looks for a sheet with time-series data: date + qty columns.
    Falls back to first numeric column pair found.
    Returns (skus, file_params).
    """
    xl = pd.ExcelFile(path)
    # Try sheet names: Data, History, Forecast, first sheet
    preferred = ["Data", "History", "Input", "Sheet1"]
    sheet_name = next((s for s in preferred if s in xl.sheet_names), xl.sheet_names[0])

    df = xl.parse(sheet_name, header=0)
    df.columns = [str(c).lower().strip() for c in df.columns]

    # Detect date and qty columns
    date_col = next((c for c in df.columns
                     if any(k in c for k in ("date", "period", "month", "calmonth", "time"))), None)
    qty_col  = next((c for c in df.columns
                     if any(k in c for k in ("qty", "quantity", "volume", "sales", "shipment",
                                              "actual", "value", "demand"))), None)

    if date_col is None:
        # Try first column that looks like dates
        for c in df.columns:
            parsed = _parse_date_col(df[c].dropna().head(5))
            if any(isinstance(d, datetime) for d in parsed):
                date_col = c
                break

    if qty_col is None:
        # Take first numeric column that isn't the date col
        for c in df.columns:
            if c == date_col:
                continue
            if pd.to_numeric(df[c], errors="coerce").notna().sum() > 5:
                qty_col = c
                break

    if date_col is None or qty_col is None:
        raise ValueError(
            f"Cannot detect date/qty columns in {sheet_name}. Columns: {list(df.columns)}"
        )

    df[date_col] = _parse_date_col(df[date_col])
    df.dropna(subset=[date_col, qty_col], inplace=True)
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df = df.sort_values(date_col)

    # Try to get SKU name from sheet or filename
    sku_name = xl.sheet_names[0] if xl.sheet_names[0] not in preferred else path.split("/")[-1].replace(".xlsm", "")

    return [{
        "name":    sku_name,
        "history": df[qty_col].tolist(),
        "dates":   df[date_col].tolist(),
    }], _default_file_params()


def read_multi_sku_excel(path: str):
    """
    Read multi-SKU xlsx.
    Supports 2 layouts:
      (A) Wide table: each SKU a column, first column = date
      (B) Long table: columns date, sku, qty
      (C) Multi-sheet: each sheet = one SKU
    Returns (skus, file_params).
    """
    xl = pd.ExcelFile(path)
    skus_out = []

    for sheet_name in xl.sheet_names:
        try:
            df = xl.parse(sheet_name, header=0)
        except Exception:
            continue
        if df.empty or df.shape[1] < 2:
            continue

        df_cols = [str(c).lower().strip() for c in df.columns]

        # ── Layout B: long table ──
        date_col = next((df.columns[i] for i, c in enumerate(df_cols)
                         if any(k in c for k in ("date", "period", "month", "calmonth"))), None)
        qty_col  = next((df.columns[i] for i, c in enumerate(df_cols)
                         if any(k in c for k in ("qty", "quantity", "volume", "sales",
                                                   "shipment", "actual", "demand"))), None)
        sku_col  = next((df.columns[i] for i, c in enumerate(df_cols)
                         if any(k in c for k in ("sku", "material", "product", "item"))), None)

        if date_col is not None and qty_col is not None and sku_col is not None:
            # Long format
            df[date_col] = _parse_date_col(df[date_col])
            df.dropna(subset=[date_col, qty_col], inplace=True)
            df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
            for sku_name, grp in df.groupby(sku_col):
                grp = grp.sort_values(date_col)
                skus_out.append({
                    "name":    str(sku_name),
                    "history": grp[qty_col].tolist(),
                    "dates":   grp[date_col].tolist(),
                })
            break  # found a long-format sheet — done

        elif date_col is not None:
            # ── Layout A: wide table (each column = SKU) ──
            df[date_col] = _parse_date_col(df[date_col])
            df.dropna(subset=[date_col], inplace=True)
            df = df.sort_values(date_col)
            for col in df.columns:
                if col == date_col:
                    continue
                numeric = pd.to_numeric(df[col], errors="coerce")
                if numeric.notna().sum() < 3:
                    continue
                skus_out.append({
                    "name":    str(col),
                    "history": numeric.fillna(0).tolist(),
                    "dates":   df[date_col].tolist(),
                })
            if skus_out:
                break

    # ── Layout C: each sheet = one SKU (fallback) ──
    if not skus_out:
        for sheet_name in xl.sheet_names:
            try:
                df = xl.parse(sheet_name, header=0)
            except Exception:
                continue
            df.columns = [str(c).lower().strip() for c in df.columns]
            date_col = next((c for c in df.columns
                             if any(k in c for k in ("date", "period", "month"))), None)
            qty_col  = next((c for c in df.columns
                             if any(k in c for k in ("qty", "volume", "sales", "actual"))), None)
            if date_col and qty_col:
                df[date_col] = _parse_date_col(df[date_col])
                df.dropna(subset=[date_col, qty_col], inplace=True)
                df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
                df = df.sort_values(date_col)
                skus_out.append({
                    "name":    sheet_name,
                    "history": df[qty_col].tolist(),
                    "dates":   df[date_col].tolist(),
                })

    if not skus_out:
        raise ValueError(
            "Cannot parse Excel file. Expected: wide-table, long-table, or multi-sheet layout "
            "with date + qty columns."
        )

    return skus_out, _default_file_params()


# ══════════════════════════════════════════════════════════════
# BBL PPG READER — Mondelez Vietnam AFC format
# ══════════════════════════════════════════════════════════════

def read_bbl_ppg(
    path: str,
    history_start: datetime = None,
    history_end: datetime   = None,
    channels: list          = None,
    sheet_name: str         = "BBL PPG",
) -> tuple:
    """
    Read historical data from BBL PPG sheet (Mondelez Vietnam AFC format).

    Layout:
        - Row 19 (0-indexed) = column headers
        - Data from row 20 onward
        - KF col (11): filter 'Base Forecast'; fallback to 'Innovation' for NPD gaps
        - SI values: cols 12-23 (Jan-Dec), unit = SI (standard index volume)
        - Channel col (6): filter GT + MT by default
        - PPG col (7), Year col (4)

    Business rules applied:
        - SWEET SKUs: discontinued → use Innovation KF as history source
        - Y.RICE SKUs: NPD → use Innovation KF to fill months missing from Base Forecast
        - Base Forecast always takes priority over Innovation when both exist
        - Months with vol = 0 are kept as-is (genuine zero shipment periods)

    Args:
        path:           Path to .xlsx file
        history_start:  Start of history window (datetime). Default = 36M before history_end
        history_end:    End of history window (datetime). Default = Apr of current year
        channels:       List of channels to include. Default = ['GT', 'MT']
        sheet_name:     Sheet to read. Default = 'BBL PPG'

    Returns:
        skus        : list of { name, history, dates, channel, ppg, source_kf_info }
        file_params : default param ranges dict
        meta        : dict with window info, coverage summary
    """
    if channels is None:
        channels = ["GT", "MT"]

    # ── Default window: 36M ending Apr of current year ───────
    from dateutil.relativedelta import relativedelta
    today = datetime.today()
    if history_end is None:
        # Last full month = current month - 1, but user context = Apr 2025
        history_end = datetime(today.year, today.month, 1) - relativedelta(months=1)
        history_end = history_end.replace(day=1)
    if history_start is None:
        history_start = history_end - relativedelta(months=35)
        history_start = history_start.replace(day=1)

    n_window = (history_end.year - history_start.year) * 12 \
               + (history_end.month - history_start.month) + 1

    # ── Load sheet ────────────────────────────────────────────
    df = pd.read_excel(path, sheet_name=sheet_name, header=None)

    # Col positions (fixed layout, confirmed from AFC_P5 structure)
    KF_COL   = 11
    CHAN_COL  = 6
    PPG_COL  = 7
    YEAR_COL = 4
    SI_COLS  = list(range(12, 24))  # Jan-Dec SI values
    MONTHS   = ["Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]

    # Data starts at row 20 (0-indexed)
    data = df.iloc[20:].copy().reset_index(drop=True)

    # ── Extract long format for given KF types ────────────────
    def _extract(kf_values):
        mask = (data[KF_COL].isin(kf_values)) & (data[CHAN_COL].isin(channels))
        sub  = data[mask].copy()
        sub["_year"] = sub[YEAR_COL].astype(str).str.strip()
        sub = sub[sub["_year"].str.match(r"^\d{4}$")]
        sub["_chan"] = sub[CHAN_COL].astype(str)
        sub["_ppg"]  = sub[PPG_COL].astype(str)
        for i, m in enumerate(MONTHS):
            sub[m] = pd.to_numeric(sub[SI_COLS[i]], errors="coerce").fillna(0)
        rows = []
        for _, row in sub.iterrows():
            for m in MONTHS:
                month_num = MONTHS.index(m) + 1
                dt = datetime(int(row["_year"]), month_num, 1)
                rows.append({
                    "date":    dt,
                    "channel": row["_chan"],
                    "ppg":     row["_ppg"],
                    "vol":     float(row[m]),
                    "kf":      row[KF_COL],
                })
        return pd.DataFrame(rows) if rows else pd.DataFrame(
            columns=["date","channel","ppg","vol","kf"])

    bf_long    = _extract(["Base Forecast"])
    innov_long = _extract(["Innovation"])

    # ── Merge: Base Forecast primary, Innovation fills gaps ───
    # Rule: use Base Forecast when vol > 0; otherwise fall back to Innovation.
    # This handles: SWEET (discontinued, BF has zeros pre-2025),
    # Y.RICE (NPD, BF has zeros before launch date).
    bf_long["priority"]    = 0
    innov_long["priority"] = 1

    combined = pd.concat([bf_long, innov_long], ignore_index=True)
    # Sort: Base Forecast first, then Innovation
    combined = combined.sort_values(["date","channel","ppg","priority"])
    # Keep Base Forecast only if vol > 0; else keep Innovation
    combined["_use"] = False
    for idx, grp in combined.groupby(["date","channel","ppg"]):
        bf_rows    = grp[grp["priority"] == 0]
        innov_rows = grp[grp["priority"] == 1]
        bf_has_val = not bf_rows.empty and (bf_rows["vol"] > 0).any()
        if bf_has_val:
            combined.loc[bf_rows.index[0], "_use"] = True
        elif not innov_rows.empty:
            combined.loc[innov_rows.index[0], "_use"] = True
        elif not bf_rows.empty:
            # Base Forecast exists but vol=0 → keep it as genuine zero
            combined.loc[bf_rows.index[0], "_use"] = True
    combined = combined[combined["_use"]].drop(columns=["priority","_use"])

    # ── Filter to history window ──────────────────────────────
    window = combined[
        (combined["date"] >= history_start) &
        (combined["date"] <= history_end)
    ].copy()

    # ── Build SKU list (PPG x Channel) ───────────────────────
    skus_out    = []
    coverage    = {}

    all_ppg_chan = (
        window[window["vol"] > 0]
        .groupby(["ppg","channel"])["date"]
        .count()
        .reset_index()
    )
    all_ppg_chan = all_ppg_chan[all_ppg_chan["date"] > 0]

    # Build full date range for the window
    full_dates = pd.date_range(history_start, history_end, freq="MS").to_pydatetime().tolist()

    for _, pc_row in all_ppg_chan.iterrows():
        ppg_name = pc_row["ppg"]
        channel  = pc_row["channel"]

        # Get all months in window for this series
        series_df = window[
            (window["ppg"] == ppg_name) &
            (window["channel"] == channel)
        ].set_index("date")["vol"].reindex(full_dates, fill_value=0)

        history_vals = series_df.values.tolist()
        hist_dates   = list(full_dates)

        # Track which KF source was used
        source_info = window[
            (window["ppg"] == ppg_name) &
            (window["channel"] == channel) &
            (window["vol"] > 0)
        ]["kf"].value_counts().to_dict()

        n_nonzero  = sum(1 for v in history_vals if v > 0)
        n_zero     = n_window - n_nonzero

        skus_out.append({
            "name":          f"{ppg_name} | {channel}",
            "history":       history_vals,
            "dates":         hist_dates,
            "channel":       channel,
            "ppg":           ppg_name,
            "source_kf":     source_info,
            "n_months":      n_window,
            "n_nonzero":     n_nonzero,
            "n_zero_filled": n_zero,
        })

        coverage[f"{ppg_name} | {channel}"] = {
            "n_months":      n_window,
            "n_nonzero":     n_nonzero,
            "n_zero_filled": n_zero,
            "source_kf":     source_info,
        }

    if not skus_out:
        raise ValueError(
            f"No data found in BBL PPG for channels={channels}, "
            f"window={history_start.strftime('%b %Y')} → {history_end.strftime('%b %Y')}. "
            "Check KF='Base Forecast' or 'Innovation' rows exist."
        )

    # ── Pull Jan-Apr 2023 for full Y2023 YoY anchor calc ────
    # These months are OUTSIDE the 36M history window but needed
    # to compute a complete Y2023 annual total for YoY anchor.
    yoy_anchor_extra_all = {}   # keyed by (ppg, channel)
    try:
        anchor_start = datetime(2023, 1, 1)
        anchor_end   = datetime(2023, 4, 1)
        if history_start > anchor_start:   # only pull if not already in window
            anchor_window = combined[
                (combined["date"] >= anchor_start) &
                (combined["date"] <= anchor_end)
            ].copy()
            for (ppg_n, chan_n), grp in anchor_window.groupby(["ppg", "channel"]):
                grp = grp.sort_values("date")
                yoy_anchor_extra_all[(ppg_n, chan_n)] = [
                    (row["date"], row["vol"])
                    for _, row in grp.iterrows() if row["vol"] > 0
                ]
    except Exception:
        pass

    # Attach yoy_anchor_extra to each SKU
    for s in skus_out:
        key = (s["ppg"], s["channel"])
        s["yoy_anchor_extra"] = yoy_anchor_extra_all.get(key, [])

    meta = {
        "sheet":           sheet_name,
        "history_start":   history_start,
        "history_end":     history_end,
        "n_window_months": n_window,
        "channels":        channels,
        "n_skus":          len(skus_out),
        "coverage":        coverage,
    }

    return skus_out, _default_file_params(), meta


# ══════════════════════════════════════════════════════════════
# EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════

def export_to_excel(path: str, all_sku_results: list):
    """
    Export forecast results to Excel.
    all_sku_results: list of {
        name, history, hist_dates, fc_dates, top (list of result dicts)
    }
    Sheets: Summary, Results, Growth Analysis
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: pandas ExcelWriter
        _export_pandas_fallback(path, all_sku_results)
        return

    wb = Workbook()

    # ── Sheet: Summary ────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    hdr_fill  = PatternFill("solid", fgColor="1E2130")
    hdr_font  = Font(name="Arial", bold=True, color="4A9EFF", size=10)
    val_font  = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)

    def _hdr(ws, row, col, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    def _val(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = val_font
        cell.alignment = Alignment(horizontal="center")
        return cell

    sum_headers = ["SKU", "Best Model", "Alpha", "Beta", "Gamma",
                   "MAPE", "Score", "H-Steps"]
    for ci, h in enumerate(sum_headers, 1):
        _hdr(ws_sum, 1, ci, h)

    for ri, sku_res in enumerate(all_sku_results, 2):
        best = sku_res["top"][0]
        from config import MODEL_LABELS
        _val(ws_sum, ri, 1, sku_res["name"])
        _val(ws_sum, ri, 2, MODEL_LABELS.get(best["model"], best["model"]))
        _val(ws_sum, ri, 3, best["alpha"])
        _val(ws_sum, ri, 4, best["beta"])
        _val(ws_sum, ri, 5, best["gamma"])
        mape_v = best["mape"]
        _val(ws_sum, ri, 6, f"{mape_v*100:.1f}%" if not math.isnan(mape_v) else "N/A")
        _val(ws_sum, ri, 7, round(best["score"], 5))
        _val(ws_sum, ri, 8, len(sku_res["fc_dates"]))

    for col in ws_sum.columns:
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = 16

    # ── Sheet: Results ────────────────────────────────────────
    ws_res = wb.create_sheet("Results")
    res_headers = ["SKU", "Model", "Rank", "Alpha", "Beta", "Gamma", "MAPE", "Score"]

    for sku_res in all_sku_results:
        fc_dates = sku_res["fc_dates"]
        res_headers_full = res_headers + [d.strftime("%b %Y") for d in fc_dates]
        for ci, h in enumerate(res_headers_full, 1):
            _hdr(ws_res, 1, ci, h)
        break

    row_idx = 2
    for sku_res in all_sku_results:
        for rank, r in enumerate(sku_res["top"], 1):
            mape_v = r["mape"]
            base_vals = [
                sku_res["name"],
                r["model"],
                rank,
                r["alpha"],
                r["beta"],
                r["gamma"],
                f"{mape_v*100:.1f}%" if not math.isnan(mape_v) else "N/A",
                round(r["score"], 5),
            ]
            for ci, v in enumerate(base_vals, 1):
                _val(ws_res, row_idx, ci, v)
            for ci, v in enumerate(r["forecast"], len(base_vals) + 1):
                _val(ws_res, row_idx, ci, round(v, 1))
            row_idx += 1

    for col in ws_res.columns:
        ws_res.column_dimensions[get_column_letter(col[0].column)].width = 13

    # ── Sheet: Growth Analysis ────────────────────────────────
    ws_growth = wb.create_sheet("Growth Analysis")
    ga_headers = ["SKU", "Period", "Type", "Volume", "YoY %", "QoQ %", "MoM %"]
    for ci, h in enumerate(ga_headers, 1):
        _hdr(ws_growth, 1, ci, h)

    ga_row = 2
    for sku_res in all_sku_results:
        best  = sku_res["top"][0]
        hist  = sku_res["history"]
        hd    = sku_res["hist_dates"]
        fc    = best["forecast"]
        fcd   = sku_res["fc_dates"]

        all_v  = hist + fc
        all_d  = hd + fcd
        last_h = hd[-1]

        for i, (dt, v) in enumerate(zip(all_d, all_v)):
            typ    = "History" if dt <= last_h else "Forecast"
            mom    = (v / all_v[i-1] - 1) if i > 0 and all_v[i-1] != 0 else None
            yoy    = (v / all_v[i-12] - 1) if i >= 12 and all_v[i-12] != 0 else None
            # QoQ: compare to 3 months ago (rough)
            qoq    = (v / all_v[i-3] - 1)  if i >= 3  and all_v[i-3]  != 0 else None

            def _pct(val):
                return f"{val*100:+.1f}%" if val is not None else ""

            _val(ws_growth, ga_row, 1, sku_res["name"])
            _val(ws_growth, ga_row, 2, dt.strftime("%b %Y"))
            _val(ws_growth, ga_row, 3, typ)
            _val(ws_growth, ga_row, 4, round(v, 1))
            _val(ws_growth, ga_row, 5, _pct(yoy))
            _val(ws_growth, ga_row, 6, _pct(qoq))
            _val(ws_growth, ga_row, 7, _pct(mom))
            ga_row += 1

    for col in ws_growth.columns:
        ws_growth.column_dimensions[get_column_letter(col[0].column)].width = 14

    wb.save(path)


def _export_pandas_fallback(path: str, all_sku_results: list):
    """Pandas fallback if openpyxl unavailable."""
    rows = []
    for sku_res in all_sku_results:
        best = sku_res["top"][0]
        mape_v = best["mape"]
        rows.append({
            "SKU":   sku_res["name"],
            "Model": best["model"],
            "Alpha": best["alpha"],
            "Beta":  best["beta"],
            "Gamma": best["gamma"],
            "MAPE":  f"{mape_v*100:.1f}%" if not math.isnan(mape_v) else "N/A",
            "Score": round(best["score"], 5),
        })
    pd.DataFrame(rows).to_excel(path, index=False)